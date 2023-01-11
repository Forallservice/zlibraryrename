import openpyxl
import os


def rxl(name):

    global workbook 
    table = workbook.active
    rows = table.max_row

    for row in range(rows):
    	if str(table.cell(row + 1, 1).value)==name:		
    		return str(table.cell(row + 1, 6).value)+"."+str(table.cell(row + 1, 4).value)

#修改下面的文件名为改名的数据库文件，有时可能需要两个才能完全覆盖一个文件包
input_file_name = '21000001-21500000.xlsx'
global workbook
workbook  = openpyxl.load_workbook(input_file_name)
#修改下面的路径为你需要批量改名的目录
dst_path = 'd:\zlib\pilimi-zlib2-21230000-21319999'
ext_name = ''
for r, d, files in os.walk( dst_path ):
    for file in files:
        if file.find(".")==-1:
            print( file,end='' )
            fn=rxl(file)
            if fn!=None:
            	print( "=>"+fn )
            	oldfile=os.path.join( dst_path ,file)
            	fn=fn.replace("?","")
            	fn=fn.replace(":","")
            	fn=fn.replace("\\","")
            	fn=fn.replace("/"," ")
            	fn=fn.replace('"'," ")
            	fn=fn.replace("*"," ")
            	fn=fn.replace('<'," ")
            	fn=fn.replace(">"," ")
            	fn=fn.replace("|"," ")
            	fn=fn.replace("LPT1.",file+"LPT1.")
            	fn=fn[-125:]
            	#print(fn)
            	newfile=os.path.join( dst_path ,fn)
            	newfile2=os.path.join( dst_path ,file)+fn
            	if os.path.exists(newfile):
            		os.rename(oldfile, newfile2)
            	else:
            		os.rename(oldfile,newfile)


